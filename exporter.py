"""
Excel エクスポートモジュール
"""

from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COLUMNS = [
    ("法人番号",           "法人番号",           20),
    ("事業所番号",         "事業所番号",         22),
    ("事業所名",           "事業所名",           30),
    ("事業所ホームページ", "事業所ホームページ", 40),
    ("産業分類",           "産業分類",           25),
    ("従業員数（企業全体）", "従業員数企業全体",  18),
    ("従業員数（就業場所）", "従業員数就業場所",  18),
    ("資本金",             "資本金",             15),
    ("創業年",             "創業年",             12),
    ("事業内容",           "事業内容",           50),
    ("会社の特長",         "会社の特長",         50),
    ("代表者役職",         "代表者役職",         18),
    ("代表者名",           "代表者名",           18),
    ("就業場所住所",       "就業場所住所",       40),
    ("転勤範囲",           "転勤範囲",           20),
    ("事業所所在地",       "事業所所在地",       40),
    ("職種",               "職種",               25),
    ("仕事内容",           "仕事内容",           60),
    ("必要な経験等",       "必要な経験等",       40),
    ("必要なPCスキル",     "必要なPCスキル",     30),
    ("必要な免許・資格",   "必要な免許・資格",   30),
    ("採用人数",           "採用人数",           12),
    ("募集理由区分",       "募集理由区分",       20),
    ("その他募集理由",     "その他募集理由",     40),
    ("求人に関する特記事項", "求人に関する特記事項", 60),
    ("事業内容・会社の特長", "事業内容・会社の特長", 50),
    ("求人番号",           "求人番号",           25),
    ("詳細URL",            "詳細URL",            60),
]

HEADER_FILL_CORP = PatternFill(fill_type="solid", fgColor="1F4E79")
HEADER_FILL_JOB  = PatternFill(fill_type="solid", fgColor="0070C0")
HEADER_FILL_META = PatternFill(fill_type="solid", fgColor="7F7F7F")
HEADER_FONT      = Font(bold=True, color="FFFFFF", size=10)
DATA_FONT        = Font(size=10)
BORDER_SIDE      = Side(style="thin", color="CCCCCC")
CELL_BORDER      = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
ROW_FILL_ODD     = PatternFill(fill_type="solid", fgColor="EBF3FB")
ROW_FILL_EVEN    = PatternFill(fill_type="solid", fgColor="FFFFFF")

CORP_COL_COUNT = 16
JOB_COL_COUNT  = 10


class ExcelExporter:
    def export(self, records: List[Dict], filepath: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "求人情報"
        ws.freeze_panes = "A2"

        for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = CELL_BORDER

            if col_idx <= CORP_COL_COUNT:
                cell.fill = HEADER_FILL_CORP
            elif col_idx <= CORP_COL_COUNT + JOB_COL_COUNT:
                cell.fill = HEADER_FILL_JOB
            else:
                cell.fill = HEADER_FILL_META

            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 30

        for row_idx, record in enumerate(records, start=2):
            fill = ROW_FILL_ODD if row_idx % 2 == 1 else ROW_FILL_EVEN
            for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
                value = record.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = CELL_BORDER
                cell.fill = fill

            ws.row_dimensions[row_idx].height = 60

        ws.insert_rows(1)
        corp_end = CORP_COL_COUNT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=corp_end)
        c = ws.cell(row=1, column=1, value="【法人情報】")
        c.fill = HEADER_FILL_CORP
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")

        job_start = CORP_COL_COUNT + 1
        job_end = CORP_COL_COUNT + JOB_COL_COUNT
        ws.merge_cells(start_row=1, start_column=job_start, end_row=1, end_column=job_end)
        c = ws.cell(row=1, column=job_start, value="【求人情報】")
        c.fill = HEADER_FILL_JOB
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")

        meta_start = CORP_COL_COUNT + JOB_COL_COUNT + 1
        meta_end = len(COLUMNS)
        ws.merge_cells(start_row=1, start_column=meta_start, end_row=1, end_column=meta_end)
        c = ws.cell(row=1, column=meta_start, value="【管理情報】")
        c.fill = HEADER_FILL_META
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 22

        wb.save(filepath)
